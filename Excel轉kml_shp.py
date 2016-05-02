# -*- coding:utf8  -*-
'''
程式名稱：作業10程式
版本：1.0
作者：S.W.
說明：讀取Excel檔案並轉出kml檔和shapefile檔。
'''

#載入模組
from openpyxl import load_workbook
import shapefile
import twd97
from math import radians

#-----------------------------------------

#轉出Kml檔
def Output_kml(fileIn, fileOut):

    #讀取Excel檔案
    wb = load_workbook(fileIn)
    sheets = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheets[0])
    
    #建立kml檔案
    fout = open(fileOut, 'w')
    
    #新增kml內容
    fout.write('<?xml version="1.0" encoding="UTF-8"?>\n')
    fout.write('<kml>\n')
    fout.write('<Folder>\n')
    rows = ws.rows
    for i in range(len(rows)):
        if i >= 1 :
            fout.write('<Placemark>\n')
            fout.write('    <name>' + ws.cell(row = i+1, column = 1).value.encode('utf8') + '</name>\n')
            fout.write('    <description>這是 ' + ws.cell(row = i+1, column = 1).value.encode('utf8') + '</description>\n')
            fout.write('    <Point>\n')
            fout.write('        <coordinates>' + str(ws.cell(row = i+1, column = 3).value) + ',' + str(ws.cell(row = i+1, column = 2).value) + '</coordinates>\n')
            fout.write('    </Point>\n')
            fout.write('</Placemark>\n')
            fout.write('\n')
            fout.write('\n')
        else:
            None    
    fout.write('</Folder>\n')   
    fout.write('</kml>\n')
    fout.close()
    print "kml檔 %s 產製成功!" % fileOut

#-----------------------------------------

#轉出Shapefile檔
def Output_shp(filein, fileOut):

    #讀取Excel檔案
    wb = load_workbook(filein)
    sheets = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheets[0])

    #建立Shapefile檔案
    shp = shapefile.Writer(shapefile.POINT)

    #新增屬性欄位
    shp.field('Name','C','10',)
    shp.field('Lat','F','15',3)
    shp.field('Lon','F','15',3)
    shp.field('URL','C','50')
    shp.field('Remarks','C','50')
    
    #加入空間資料和屬性資料
    rows = ws.rows
    for i in range(len(rows)):
        if i >= 1 :
            Name = ws.cell(row = i+1, column = 1).value.encode('big5')
            Lat = (ws.cell(row = i+1, column = 2).value)
            Lon = (ws.cell(row = i+1, column = 3).value)
            URL = ws.cell(row = i+1, column = 4).value.encode('big5')
            Remarks = ws.cell(row = i+1, column = 5).value.encode('big5')
            t = twd97.LatLonToTWD97()
            X, Y = t.convert(radians(Lat), radians(Lon))
          
            shp.point(X,Y) 
            shp.record(Name,Lat,Lon,URL,Remarks) 

    #儲存Shapefile檔案
    try:
        shp.save(fileOut)
        print "Shapefile圖檔 %s 產製成功!" % fileOut
    except:
        print "Shapefile圖檔 %s 產製失敗!" % fileOut
    
#-----------------------------------------

if __name__ == '__main__':
    print '資料選取(不須加副檔名)'
    fin = raw_input('請輸入Excel檔名:') + '.xlsx'
    fout1 = raw_input('請輸入kml檔名:') + '.kml'
    fout2 = raw_input('請輸入Shapefile檔名:')
    Output_kml(fin, fout1)
    Output_shp(fin, fout2)
    
